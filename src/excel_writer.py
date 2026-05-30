from __future__ import annotations

import io
import re
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .table_extractor import ExtractedTable
from .utils import clean_filename, clean_sheet_name


HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True)


def _fallback_sheet_name(table: ExtractedTable, used: set[str]) -> str:
    first_text = ""
    for row in table.rows[:2]:
        for cell in row:
            if cell:
                first_text = cell
                break
        if first_text:
            break
    base = clean_sheet_name(f"P{table.page}_{first_text[:18]}", fallback=f"P{table.page}_{table.index}")
    name = base
    counter = 2
    while name in used:
        suffix = f"_{counter}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        counter += 1
    used.add(name)
    return name


def _ai_sheet_name(table: ExtractedTable, claude_api_key: str) -> str:
    if not claude_api_key:
        return ""
    try:
        import anthropic

        client = anthropic.Anthropic(api_key=claude_api_key)
        sample = table.preview_text
        message = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=80,
            messages=[
                {
                    "role": "user",
                    "content": (
                        "请为这个财报表格生成一个简短 Excel Sheet 名称，12 个中文字符以内，"
                        "不要解释，不要标点，不要特殊符号。表格内容："
                        f"{sample}"
                    ),
                }
            ],
        )
        text = "".join(block.text for block in message.content if getattr(block, "type", "") == "text").strip()
        text = re.sub(r"[\s\[\]:*?/\\]+", "", text)
        return text[:12]
    except Exception:
        return ""


def _style_sheet(sheet) -> None:
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column in sheet.columns:
        max_length = 8
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value) + 2, 50))
        sheet.column_dimensions[col_letter].width = max_length
    sheet.freeze_panes = "A2"


def tables_to_workbook_bytes(tables: Iterable[ExtractedTable], claude_api_key: str = "") -> bytes:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    used_names: set[str] = set()
    count = 0
    for table in tables:
        ai_name = _ai_sheet_name(table, claude_api_key)
        if ai_name:
            base_name = clean_sheet_name(ai_name, fallback=f"P{table.page}_{table.index}")
            name = base_name
            counter = 2
            while name in used_names:
                suffix = f"_{counter}"
                name = f"{base_name[:31 - len(suffix)]}{suffix}"
                counter += 1
            used_names.add(name)
        else:
            name = _fallback_sheet_name(table, used_names)
        sheet = workbook.create_sheet(title=name)
        for row in table.rows:
            sheet.append(row)
        _style_sheet(sheet)
        count += 1
    if count == 0:
        sheet = workbook.create_sheet(title="无可提取表格")
        sheet.append(["未从 PDF 中提取到文字版表格"])
        _style_sheet(sheet)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def excel_filename(pdf_name: str) -> str:
    stem = re.sub(r"\.pdf$", "", pdf_name or "", flags=re.I)
    return f"{clean_filename(stem, 'pdf')}_tables.xlsx"
